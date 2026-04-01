#!/usr/bin/env python3
"""
import_reference_data.py — Sync reference data to the de.NBI Service Registry admin API
========================================================================================

Reads an Excel workbook with three sheets and upserts each row via the admin REST API.
The script never deletes anything; missing optional fields are skipped gracefully.

Sheets expected (exact names):
  - pis              → /api/v1/pis/
  - service_categories → /api/v1/categories/
  - service_centers  → /api/v1/service-centers/

Supported column headers per sheet
───────────────────────────────────
pis
  last_name (required)  first_name (required)  email  institute  orcid
  is_active (true/false)  is_associated_partner (true/false)

service_categories
  name (required)  is_active (true/false)

service_centers
  short_name (required)  full_name (required)  website  is_active (true/false)

Boolean columns accept:  true / false / yes / no / 1 / 0  (case-insensitive)
Blank boolean cells fall back to the API default (True for is_active, False for
is_associated_partner).

Update logic
────────────
  • Each row is matched against existing records fetched at startup (all statuses).
  • Match key: pis → (last_name, first_name)  |  categories → name  |  centers → short_name
  • No match  → POST  (create)
  • Match found, values differ  → PATCH  (partial update)
  • Match found, values identical  → skip  (no request sent)
  • Records that exist in the API but are absent from the sheet are left untouched.

Usage
─────
  python import_reference_data.py --api-url https://your-registry.example.com \\
                                   --api-key  YOUR_ADMIN_TOKEN \\
                                   --file     backend_admin_template_filled.xlsx

  python import_reference_data.py --help

Options
───────
  --api-url   Base URL of the registry (no trailing slash).  [required]
  --api-key   Admin token (Authorization: Token <key>).       [required]
  --file      Path to the .xlsx file.                         [required]
  --dry-run   Print planned changes without sending any request.
  --verbose   Print every row, even skipped ones.
  --timeout   Per-request timeout in seconds (default: 30).
"""

import argparse
import sys
import textwrap
import urllib.parse
from typing import Any

# ---------------------------------------------------------------------------
# Dependency guard — give a clear install hint before importing anything else
# ---------------------------------------------------------------------------
try:
    import openpyxl
except ImportError:
    sys.exit(
        "ERROR: 'openpyxl' is not installed.\n"
        "Install it with:  pip install openpyxl\n"
        "or:               conda install -c conda-forge openpyxl"
    )

try:
    import requests
except ImportError:
    sys.exit(
        "ERROR: 'requests' is not installed.\n"
        "Install it with:  pip install requests"
    )


# ---------------------------------------------------------------------------
# Sheet / endpoint configuration
# ---------------------------------------------------------------------------

# Each entry describes one sheet and its API endpoint.
# required_fields  – API will reject a POST that omits these
# optional_fields  – silently skipped when blank
# boolean_fields   – normalise to Python bool before sending
# match_key        – tuple of field names used to find existing records
SHEET_CONFIG: dict[str, dict[str, Any]] = {
    "pis": {
        "endpoint": "/api/v1/pis/",
        "required_fields": ["last_name", "first_name"],
        "optional_fields": ["email", "institute", "orcid"],
        "boolean_fields": ["is_active", "is_associated_partner"],
        "match_key": ("last_name", "first_name"),
        # Column aliases: Excel header → API field name
        "column_map": {
            "is_active (true/false)": "is_active",
            "is_associated_partner (true/false)": "is_associated_partner",
        },
    },
    "service_categories": {
        "endpoint": "/api/v1/categories/",
        "required_fields": ["name"],
        "optional_fields": [],
        "boolean_fields": ["is_active"],
        "match_key": ("name",),
        "column_map": {
            "is_active (true/false)": "is_active",
        },
    },
    "service_centers": {
        "endpoint": "/api/v1/service-centers/",
        "required_fields": ["short_name", "full_name"],
        "optional_fields": ["website"],
        "boolean_fields": ["is_active"],
        "match_key": ("short_name",),
        "column_map": {
            "is_active (true/false)": "is_active",
        },
    },
}

# Boolean values accepted from the spreadsheet
_TRUE_VALUES = {"true", "yes", "1"}
_FALSE_VALUES = {"false", "no", "0"}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _normalise_header(raw: str) -> str:
    """Strip surrounding whitespace from a column header."""
    return str(raw).strip() if raw is not None else ""


def _parse_bool(value: Any) -> bool | None:
    """
    Convert a spreadsheet cell value to Python bool.
    Returns None when the cell is blank (caller decides the fallback).
    Raises ValueError for unrecognised strings.
    """
    if value is None or (isinstance(value, str) and value.strip() == ""):
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    s = str(value).strip().lower()
    if s in _TRUE_VALUES:
        return True
    if s in _FALSE_VALUES:
        return False
    raise ValueError(
        f"Cannot interpret {value!r} as a boolean. "
        f"Use: true/false, yes/no, 1/0 (case-insensitive)."
    )


def _cell_str(value: Any) -> str:
    """Return the cell value as a stripped string, or empty string if None."""
    if value is None:
        return ""
    return str(value).strip()


def _build_url(base: str, path: str) -> str:
    """Join a base URL and a path, normalising slashes."""
    return base.rstrip("/") + "/" + path.lstrip("/")


def _records_differ(existing: dict, payload: dict) -> bool:
    """
    Return True if any key in payload has a different value in existing.
    Only compares keys present in payload (partial-update semantics).
    String comparisons are case-sensitive and whitespace-stripped.
    """
    for key, new_val in payload.items():
        old_val = existing.get(key)
        if isinstance(new_val, str):
            if _cell_str(old_val) != new_val:
                return True
        elif isinstance(new_val, bool):
            # API returns JSON booleans; Python bool comparison is exact
            if old_val != new_val:
                return True
        else:
            if old_val != new_val:
                return True
    return False


# ---------------------------------------------------------------------------
# API client
# ---------------------------------------------------------------------------

class RegistryAPIClient:
    """Thin wrapper around requests for the admin API."""

    def __init__(self, base_url: str, api_key: str, timeout: int = 30) -> None:
        self._base = base_url.rstrip("/")
        self._timeout = timeout
        self._session = requests.Session()
        self._session.headers.update({
            "Authorization": f"Token {api_key}",
            "Content-Type": "application/json",
            "Accept": "application/json",
        })

    def _url(self, path: str) -> str:
        return _build_url(self._base, path)

    def get_all(self, endpoint: str) -> list[dict]:
        """Fetch every record from an endpoint (pagination disabled server-side)."""
        url = self._url(endpoint)
        resp = self._session.get(url, timeout=self._timeout)
        _raise_for_status(resp, "GET", url)
        data = resp.json()
        # Handle both plain list and DRF paginated envelope
        if isinstance(data, list):
            return data
        return data.get("results", data)

    def create(self, endpoint: str, payload: dict) -> dict:
        url = self._url(endpoint)
        resp = self._session.post(url, json=payload, timeout=self._timeout)
        _raise_for_status(resp, "POST", url, payload)
        return resp.json()

    def update(self, endpoint: str, record_id: Any, payload: dict) -> dict:
        url = self._url(endpoint) + f"{record_id}/"
        resp = self._session.patch(url, json=payload, timeout=self._timeout)
        _raise_for_status(resp, "PATCH", url, payload)
        return resp.json()

    def verify_auth(self, endpoint: str) -> None:
        """
        Do a single GET to confirm the token is accepted.
        Raises SystemExit on 401/403.
        """
        url = self._url(endpoint)
        resp = self._session.get(url, timeout=self._timeout)
        if resp.status_code in (401, 403):
            sys.exit(
                f"ERROR: Authentication failed (HTTP {resp.status_code}).\n"
                "Check that --api-key is a valid admin token."
            )


def _raise_for_status(
    resp: "requests.Response",
    method: str,
    url: str,
    payload: dict | None = None,
) -> None:
    """Raise a RuntimeError with context when the API returns an error status."""
    if resp.ok:
        return
    body = resp.text[:400]
    payload_hint = f"\n  Payload: {payload}" if payload else ""
    raise RuntimeError(
        f"API error {resp.status_code} on {method} {url}{payload_hint}\n"
        f"  Response: {body}"
    )


# ---------------------------------------------------------------------------
# Excel reader
# ---------------------------------------------------------------------------

def _read_sheet(
    ws: "openpyxl.worksheet.worksheet.Worksheet",
    config: dict,
    sheet_name: str,
) -> list[dict]:
    """
    Parse one worksheet into a list of field-value dicts ready for the API.
    Validates required fields and normalises booleans.
    Rows that are completely empty are silently skipped.
    Returns (rows, errors) — errors is a list of human-readable strings.
    """
    column_map: dict[str, str] = config["column_map"]
    required_fields: list[str] = config["required_fields"]
    optional_fields: list[str] = config["optional_fields"]
    boolean_fields: list[str] = config["boolean_fields"]

    # Read header row
    raw_headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    if all(h is None for h in raw_headers):
        raise ValueError(f"Sheet '{sheet_name}': header row (row 1) is empty.")

    # Map column index → canonical API field name
    col_to_field: dict[int, str] = {}
    for col_idx, raw in enumerate(raw_headers, start=1):
        header = _normalise_header(raw)
        if not header:
            continue
        # Resolve alias first, then use as-is
        field_name = column_map.get(header, header)
        col_to_field[col_idx] = field_name

    known_fields = set(required_fields) | set(optional_fields) | set(boolean_fields)
    unknown = {
        f for f in col_to_field.values() if f not in known_fields
    }
    if unknown:
        print(
            f"  [WARN] Sheet '{sheet_name}': unrecognised column(s) will be ignored: "
            + ", ".join(sorted(unknown))
        )

    rows: list[dict] = []
    errors: list[str] = []

    for row_idx in range(2, ws.max_row + 1):
        raw_row = {
            col_to_field[c]: ws.cell(row_idx, c).value
            for c in col_to_field
        }

        # Skip entirely blank rows
        if all(v is None or _cell_str(v) == "" for v in raw_row.values()):
            continue

        payload: dict[str, Any] = {}
        row_errors: list[str] = []

        # Required string fields
        for field in required_fields:
            val = _cell_str(raw_row.get(field, ""))
            if not val:
                row_errors.append(
                    f"sheet='{sheet_name}' row={row_idx}: "
                    f"required field '{field}' is empty — row skipped."
                )
            else:
                payload[field] = val

        if row_errors:
            errors.extend(row_errors)
            continue  # skip the row; required field missing

        # Optional string fields — include only when non-blank
        for field in optional_fields:
            val = _cell_str(raw_row.get(field, ""))
            if val:
                payload[field] = val

        # Boolean fields — include only when the cell has a value
        for field in boolean_fields:
            raw_bool = raw_row.get(field)
            try:
                parsed = _parse_bool(raw_bool)
            except ValueError as exc:
                errors.append(
                    f"sheet='{sheet_name}' row={row_idx} field='{field}': {exc} — "
                    "field skipped, API default will be used."
                )
                parsed = None
            if parsed is not None:
                payload[field] = parsed

        rows.append(payload)

    return rows, errors


# ---------------------------------------------------------------------------
# Sync logic per sheet
# ---------------------------------------------------------------------------

def _sync_sheet(
    client: RegistryAPIClient,
    sheet_name: str,
    config: dict,
    rows: list[dict],
    *,
    dry_run: bool,
    verbose: bool,
) -> dict[str, int]:
    """
    Upsert all rows from the sheet against the live API.
    Returns a stats dict with keys: created, updated, skipped, errors.
    """
    endpoint = config["endpoint"]
    match_key: tuple[str, ...] = config["match_key"]

    stats = {"created": 0, "updated": 0, "skipped": 0, "errors": 0}

    # Fetch current state from API
    print(f"\n  Fetching existing records from {endpoint} …")
    existing_records: list[dict] = client.get_all(endpoint)
    print(f"  Found {len(existing_records)} existing record(s).")

    # Build lookup index: match_key_tuple → record dict
    index: dict[tuple, dict] = {}
    for rec in existing_records:
        key = tuple(
            _cell_str(rec.get(k, "")).strip().casefold()
            for k in match_key
        )
        index[key] = rec

    for payload in rows:
        match_tuple = tuple(
            _cell_str(payload.get(k, "")).strip().casefold()
            for k in match_key
        )
        match_label = " / ".join(
            f"{k}={payload.get(k)!r}" for k in match_key
        )

        existing = index.get(match_tuple)

        if existing is None:
            # CREATE
            if dry_run:
                print(f"    [DRY-RUN] Would CREATE: {match_label}  payload={payload}")
            else:
                try:
                    client.create(endpoint, payload)
                    print(f"    [CREATED] {match_label}")
                except RuntimeError as exc:
                    print(f"    [ERROR]   {match_label}: {exc}")
                    stats["errors"] += 1
                    continue
            stats["created"] += 1

        elif _records_differ(existing, payload):
            # UPDATE
            record_id = existing.get("id")
            if record_id is None:
                print(
                    f"    [ERROR]   Cannot update {match_label}: "
                    "existing record has no 'id' field."
                )
                stats["errors"] += 1
                continue
            if dry_run:
                print(
                    f"    [DRY-RUN] Would PATCH id={record_id}: "
                    f"{match_label}  payload={payload}"
                )
            else:
                try:
                    client.update(endpoint, record_id, payload)
                    print(f"    [UPDATED] id={record_id}  {match_label}")
                except RuntimeError as exc:
                    print(f"    [ERROR]   {match_label}: {exc}")
                    stats["errors"] += 1
                    continue
            stats["updated"] += 1

        else:
            # SKIP — already up to date
            if verbose:
                print(f"    [SKIP]    {match_label}  (no changes)")
            stats["skipped"] += 1

    return stats


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        prog="import_reference_data.py",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        description=textwrap.dedent("""\
            Sync reference data from an Excel workbook to the de.NBI Service Registry
            admin API.  Records are created or updated — nothing is ever deleted.

            The workbook must contain three sheets named exactly:
              pis, service_categories, service_centers

            Authentication uses an admin API token:
              Authorization: Token <api-key>
        """),
        epilog=textwrap.dedent("""\
            Examples:
              # Live run against the production server
              python import_reference_data.py \\
                --api-url https://service-registry.bi.denbi.de \\
                --api-key  abc123 \\
                --file     backend_admin_template_filled.xlsx

              # Dry run to preview changes
              python import_reference_data.py \\
                --api-url http://localhost:8000 \\
                --api-key  abc123 \\
                --file     backend_admin_template_filled.xlsx \\
                --dry-run  --verbose
        """),
    )
    parser.add_argument(
        "--api-url",
        required=True,
        metavar="URL",
        help="Base URL of the registry, e.g. https://service-registry.bi.denbi.de",
    )
    parser.add_argument(
        "--api-key",
        required=True,
        metavar="TOKEN",
        help="Admin API token (the value after 'Token ' in the Authorization header).",
    )
    parser.add_argument(
        "--file",
        required=True,
        metavar="PATH",
        help="Path to the Excel workbook (.xlsx).",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        default=False,
        help="Preview changes without sending any write request to the API.",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        default=False,
        help="Print every row including those that are already up to date.",
    )
    parser.add_argument(
        "--timeout",
        type=int,
        default=30,
        metavar="SECONDS",
        help="Per-request timeout in seconds (default: 30).",
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)

    # ------------------------------------------------------------------
    # Input validation
    # ------------------------------------------------------------------
    if not args.api_url.startswith(("http://", "https://")):
        sys.exit(
            "ERROR: --api-url must start with http:// or https://\n"
            f"  Got: {args.api_url}"
        )

    if not args.api_key.strip():
        sys.exit("ERROR: --api-key must not be empty.")

    # Validate URL characters to prevent header injection
    try:
        parsed = urllib.parse.urlparse(args.api_url)
        if not parsed.netloc:
            raise ValueError("No host found in URL.")
    except ValueError as exc:
        sys.exit(f"ERROR: Invalid --api-url: {exc}")

    # Reject newlines / CR in the token to prevent header injection
    if any(c in args.api_key for c in ("\n", "\r")):
        sys.exit("ERROR: --api-key contains illegal characters (newline/CR).")

    if args.timeout < 1 or args.timeout > 300:
        sys.exit("ERROR: --timeout must be between 1 and 300 seconds.")

    # ------------------------------------------------------------------
    # Load workbook
    # ------------------------------------------------------------------
    print(f"Loading workbook: {args.file}")
    try:
        wb = openpyxl.load_workbook(args.file, read_only=True, data_only=True)
    except FileNotFoundError:
        sys.exit(f"ERROR: File not found: {args.file}")
    except Exception as exc:
        sys.exit(f"ERROR: Cannot open '{args.file}': {exc}")

    missing_sheets = [s for s in SHEET_CONFIG if s not in wb.sheetnames]
    if missing_sheets:
        sys.exit(
            f"ERROR: The workbook is missing required sheet(s): "
            + ", ".join(missing_sheets)
            + f"\n  Found sheets: {wb.sheetnames}"
        )

    # ------------------------------------------------------------------
    # Authenticate
    # ------------------------------------------------------------------
    client = RegistryAPIClient(args.api_url, args.api_key, timeout=args.timeout)
    first_endpoint = next(iter(SHEET_CONFIG.values()))["endpoint"]
    print(f"Verifying admin token against {args.api_url} …")
    client.verify_auth(first_endpoint)
    print("  Token accepted.\n")

    if args.dry_run:
        print("*** DRY-RUN MODE — no changes will be written ***\n")

    # ------------------------------------------------------------------
    # Process each sheet
    # ------------------------------------------------------------------
    total_stats: dict[str, int] = {"created": 0, "updated": 0, "skipped": 0, "errors": 0}
    all_parse_errors: list[str] = []

    for sheet_name, config in SHEET_CONFIG.items():
        print(f"{'='*60}")
        print(f"Sheet: {sheet_name}  →  {config['endpoint']}")
        print(f"{'='*60}")
        ws = wb[sheet_name]

        # Parse rows from sheet
        try:
            rows, parse_errors = _read_sheet(ws, config, sheet_name)
        except ValueError as exc:
            print(f"  [ERROR] {exc}")
            all_parse_errors.append(str(exc))
            continue

        if parse_errors:
            for err in parse_errors:
                print(f"  [WARN] {err}")
            all_parse_errors.extend(parse_errors)

        print(f"  Parsed {len(rows)} data row(s) from sheet.")

        if not rows:
            print("  Nothing to sync for this sheet.")
            continue

        # Sync to API
        try:
            stats = _sync_sheet(
                client,
                sheet_name,
                config,
                rows,
                dry_run=args.dry_run,
                verbose=args.verbose,
            )
        except RuntimeError as exc:
            print(f"  [ERROR] API error while syncing '{sheet_name}': {exc}")
            all_parse_errors.append(str(exc))
            continue
        except requests.exceptions.ConnectionError as exc:
            sys.exit(f"\nERROR: Cannot reach the API at {args.api_url}.\nDetails: {exc}")
        except requests.exceptions.Timeout:
            sys.exit(
                f"\nERROR: Request timed out after {args.timeout}s. "
                "Try increasing --timeout."
            )

        for key, val in stats.items():
            total_stats[key] += val

        print(
            f"\n  Sheet summary: "
            f"created={stats['created']}  "
            f"updated={stats['updated']}  "
            f"skipped={stats['skipped']}  "
            f"errors={stats['errors']}"
        )

    wb.close()

    # ------------------------------------------------------------------
    # Final summary
    # ------------------------------------------------------------------
    print(f"\n{'='*60}")
    print("IMPORT COMPLETE")
    if args.dry_run:
        print("(DRY-RUN — no changes were written)")
    print(
        f"  Total created : {total_stats['created']}\n"
        f"  Total updated : {total_stats['updated']}\n"
        f"  Total skipped : {total_stats['skipped']}\n"
        f"  Total errors  : {total_stats['errors']}"
    )
    if all_parse_errors:
        print(f"\n  {len(all_parse_errors)} warning(s)/error(s) occurred — review output above.")

    return 1 if total_stats["errors"] or all_parse_errors else 0


if __name__ == "__main__":
    sys.exit(main())
